from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse, FileResponse, Http404
from django.contrib.auth import authenticate, login as auth_login
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from .forms import UserRegistrationForm, UploadFileForm, SendEmailForm, ContactForm, StudyForm, EmailAccountForm, EventForm, CompanyRegistrationForm, UserRegistrationForm, LoginForm, DepartmentForm, ApproveUserForm, ChangeRoleForm, CompanyForm
from .models import ProcessedFile, Contact, Study, EmailAccount, EmailDraft, SentEmail, ProcessedFileContactStatus, Event, ProcessingSession, UserProfile, Company, Department
import os
from django.conf import settings
from zipfile import ZipFile
import zipfile
import uuid
from django.db.models import Q, Count, F, FloatField, ExpressionWrapper
import pandas as pd
from django.core.files.storage import default_storage
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
import logging
from collections import defaultdict, OrderedDict
from django.core.mail import EmailMessage, get_connection, EmailMultiAlternatives
from django.contrib import messages
from django.db.models import Max
from django.forms import modelform_factory
from tinymce.widgets import TinyMCE
from django import forms
from django.utils.text import slugify
from django.utils.html import strip_tags
from django.utils.dateformat import format
import html
import random
from .templatetags.split_tags import split
from collections import defaultdict
from django.core import serializers
from django.utils.safestring import mark_safe
import json
from django.core.serializers.json import DjangoJSONEncoder
from datetime import datetime
import random
from django.utils.dateparse import parse_date
from django.urls import reverse
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from weasyprint import HTML
from django.template.loader import render_to_string 
import re
from PyPDF2 import PdfMerger
import colorsys
from django.contrib.auth.models import Group
from accounts.decorators import user_is_approved, admin_required, admin_department_required

logger = logging.getLogger(__name__)

def index(request):
    return render(request, 'index.html')

def register_choice(request):
    if request.method == 'POST':
        choice = request.POST.get('choice')
        if choice == 'company':
            return redirect('register_company')
        elif choice == 'user':
            return redirect('register_user')
    return render(request, 'registration/register_choice.html')

def register_company(request):
    if request.method == 'POST':
        form = CompanyRegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empresa registrada exitosamente. Ahora puedes registrar usuarios para esta empresa.')
            return redirect('login')
    else:
        form = CompanyRegistrationForm()
    return render(request, 'registration/register_company.html', {'form': form})

def register_user(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            # Obtener datos del formulario
            username = form.cleaned_data['username']
            email = form.cleaned_data['email']
            full_name = form.cleaned_data['full_name']
            password = form.cleaned_data['password1']
            company_name = form.cleaned_data['company_name']
            company_password = form.cleaned_data['company_password']
            department_name = form.cleaned_data['department_name']

            # Verificar si la empresa existe y la contraseña es correcta
            try:
                company = Company.objects.get(name=company_name)
                if not company.check_password(company_password):
                    form.add_error('company_password', 'Contraseña de empresa incorrecta.')
                    return render(request, 'registration/register_user.html', {'form': form})
            except Company.DoesNotExist:
                form.add_error('company_name', 'La empresa no existe.')
                return render(request, 'registration/register_user.html', {'form': form})

            # Crear usuario
            user = form.save(commit=False)
            user.email = email
            user.first_name = full_name
            user.set_password(password)
            user.save()

            # Crear o obtener el departamento
            department = None
            if department_name:
                department, created = Department.objects.get_or_create(name=department_name, company=company)

            # Determinar el rol del usuario
            # Si es el primer usuario de la empresa, asignar 'ADMIN_TOTAL' y aprobar automáticamente
            if not UserProfile.objects.filter(company=company).exists():
                role = 'ADMIN_TOTAL'
                is_approved = True
                messages.success(request, 'Registro exitoso. Has sido asignado como Administrador Total de la empresa.')
            else:
                role = 'TECNICO_DEPARTAMENTO'  # Rol predeterminado para usuarios siguientes
                is_approved = False
                messages.success(request, 'Registro exitoso. Tu cuenta está pendiente de aprobación.')

            # Crear perfil de usuario
            user_profile = UserProfile.objects.create(
                user=user,
                company=company,
                department=department,
                role=role,
                is_approved=is_approved
            )

            return redirect('login')
    else:
        form = UserRegistrationForm()
    return render(request, 'registration/register_user.html', {'form': form})

def user_login(request):
    if request.method == 'POST':
        username_or_email = request.POST.get('username')
        password = request.POST.get('password')
        remember_me = request.POST.get('remember_me')  # Checkbox para "Recordarme"

        # Intentar autenticar con username
        user = authenticate(request, username=username_or_email, password=password)

        # Si falla, intentar autenticar con email
        if user is None:
            try:
                user_obj = User.objects.get(email=username_or_email)
                user = authenticate(request, username=user_obj.username, password=password)
            except User.DoesNotExist:
                user = None

        if user is not None:
            try:
                user_profile = UserProfile.objects.get(user=user)
            except UserProfile.DoesNotExist:
                messages.error(request, 'Perfil de usuario no encontrado.')
                return render(request, 'registration/login.html')

            if user_profile.is_approved:
                auth_login(request, user)

                if not remember_me:
                    # La sesión expira al cerrar el navegador
                    request.session.set_expiry(0)
                else:
                    # La sesión expira según la configuración por defecto
                    request.session.set_expiry(None)

                messages.success(request, f'Bienvenido, {user.username}!')
                return redirect('studies')  # Redirige a 'studies' después de login exitoso
            else:
                messages.error(request, 'Tu cuenta está pendiente de aprobación.')
        else:
            messages.error(request, 'Nombre de usuario o contraseña incorrectos.')

    return render(request, 'registration/login.html')

@login_required
def dashboard(request):
    return render(request, 'dashboard.html')

# Helper functions (make sure these are included in your views.py or imported appropriately)
def get_folder_size(folder_path):
    total_size = 0
    for dirpath, dirnames, filenames in os.walk(folder_path):
        for filename in filenames:
            file_path = os.path.join(dirpath, filename)
            total_size += os.path.getsize(file_path)
    return total_size

def delete_oldest_expediente(user_folder):
    # List all subfolders (expedientes)
    expediente_folders = [(os.path.join(user_folder, folder), os.path.getmtime(os.path.join(user_folder, folder)))
                          for folder in os.listdir(user_folder)
                          if os.path.isdir(os.path.join(user_folder, folder))]
    
    if expediente_folders:
        # Sort by oldest modification date
        oldest_folder = min(expediente_folders, key=lambda x: x[1])[0]
        # Delete the oldest folder
        if os.path.exists(oldest_folder):
            os.rmdir(oldest_folder)

@login_required
@user_is_approved
def upload_file(request):
    # Obtener los estudios del usuario, ordenados por 'expediente' descendente
    studies = Study.objects.filter(user=request.user).order_by('-expediente')

    # Obtener el parámetro 'study' de la URL
    study_id = request.GET.get('study')

    if request.method == 'POST':
        # Inicializar el formulario con los datos POST y FILES, y pasar el usuario
        form = UploadFileForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            # Obtener el estudio seleccionado
            study = form.cleaned_data['study']
            study_id = study.id

            # Obtener la acción seleccionada
            action = form.cleaned_data['action']

            if action == 'process_measurements':
                # Acción: Procesar Mediciones (Excel individual)
                start_column = form.cleaned_data['start_column']
                end_column = form.cleaned_data['end_column']
                uploaded_file = form.cleaned_data['file_excel']

                # Convertir columnas de letra a índice
                try:
                    start_col_idx = column_index_from_string(start_column.upper())
                    end_col_idx = column_index_from_string(end_column.upper())
                except:
                    messages.error(request, 'Las columnas de inicio y fin deben ser letras válidas.')
                    return redirect(f"{reverse('upload_file')}?study={study_id}")

                # Procesar el archivo Excel
                try:
                    wb = load_workbook(uploaded_file)
                    ws = wb.active
                except Exception as e:
                    messages.error(request, f'Error al abrir el archivo Excel: {e}')
                    return redirect(f"{reverse('upload_file')}?study={study_id}")

                unique_values = set()
                for row in ws.iter_rows(min_row=1, min_col=start_col_idx, max_col=end_col_idx):
                    for cell in row:
                        if cell.value is not None:
                            unique_values.add(cell.value)

                if not unique_values:
                    messages.warning(request, 'No se encontraron valores únicos en el rango especificado.')
                    return redirect(f"{reverse('upload_file')}?study={study_id}")

                # Crear una nueva sesión de procesamiento
                processing_session = ProcessingSession.objects.create(user=request.user, study=study)

                # Crear la carpeta para la sesión
                user_folder = os.path.join(os.getcwd(), 'media', 'processed_files', str(request.user.id))
                expediente_folder = os.path.join(user_folder, str(study_id), str(processing_session.id))
                os.makedirs(expediente_folder, exist_ok=True)

                # Crear archivos filtrados
                for value in unique_values:
                    filtered_rows = []
                    for row in ws.iter_rows(min_row=1):
                        if any(cell.value == value for cell in row[start_col_idx-1:end_col_idx]):
                            # Excluir las columnas de gremios
                            filtered_rows.append([cell.value for idx, cell in enumerate(row) if idx < start_col_idx-1 or idx >= end_col_idx])

                    if not filtered_rows:
                        continue

                    # Crear un nuevo libro de Excel
                    new_wb = Workbook()
                    new_ws = new_wb.active

                    # Copiar filas filtradas
                    for row in filtered_rows:
                        new_ws.append(row)

                    # Copiar el ancho de columnas originales (opcional)
                    for idx, col in enumerate(ws.iter_cols(), start=1):
                        if idx < start_col_idx or idx > end_col_idx:
                            new_ws.column_dimensions[get_column_letter(idx)].width = ws.column_dimensions[get_column_letter(idx)].width

                    # Copiar la altura de filas originales (opcional)
                    for row_idx, row in enumerate(ws.iter_rows(), start=1):
                        if row_idx <= len(filtered_rows):
                            new_ws.row_dimensions[row_idx].height = ws.row_dimensions[row_idx].height

                    # Guardar el archivo
                    filename = f"{value}.xlsx"
                    file_path = os.path.join(expediente_folder, filename)
                    new_wb.save(file_path)

                    # Guardar referencia en la base de datos
                    ProcessedFile.objects.create(
                        user=request.user,
                        study=study,
                        processing_session=processing_session,
                        original_name=filename,
                        file=f'processed_files/{request.user.id}/{study_id}/{processing_session.id}/{filename}'
                    )

                # Actualizar el estado del estudio a 'Procesado'
                study.estado = 'Procesado'
                study.save()

                # Mensaje de éxito y redirección con parámetro 'study'
                messages.success(request, 'Mediciones procesadas exitosamente.')
                return redirect(f"{reverse('upload_file')}?study={study_id}")

            elif action == 'upload_packages':
                # Acción: Subir Paquetes (Archivo ZIP)
                uploaded_zip = form.cleaned_data['file_zip']

                # Crear una nueva sesión de procesamiento
                processing_session = ProcessingSession.objects.create(user=request.user, study=study)

                # Crear la carpeta para la sesión
                user_folder = os.path.join(os.getcwd(), 'media', 'processed_files', str(request.user.id))
                expediente_folder = os.path.join(user_folder, str(study_id), str(processing_session.id))
                os.makedirs(expediente_folder, exist_ok=True)

                # Descomprimir el archivo ZIP
                try:
                    with zipfile.ZipFile(uploaded_zip, 'r') as zip_ref:
                        zip_ref.extractall(expediente_folder)
                except Exception as e:
                    messages.error(request, f'Error al descomprimir el archivo ZIP: {e}')
                    return redirect(f"{reverse('upload_file')}?study={study_id}")

                # Listar los archivos extraídos
                extracted_files = os.listdir(expediente_folder)
                excel_files = [f for f in extracted_files if f.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm'))]

                if not excel_files:
                    messages.error(request, 'El archivo ZIP no contiene archivos Excel válidos.')
                    # Opcional: eliminar la carpeta vacía
                    os.rmdir(expediente_folder)
                    return redirect(f"{reverse('upload_file')}?study={study_id}")

                # Procesar cada archivo Excel extraído
                for excel_file in excel_files:
                    file_path = os.path.join(expediente_folder, excel_file)
                    try:
                        wb = load_workbook(file_path)
                        ws = wb.active
                    except Exception as e:
                        messages.error(request, f'Error al abrir el archivo Excel {excel_file}: {e}')
                        continue  # Continuar con el siguiente archivo

                    # Asumimos que los archivos ZIP ya están procesados, por lo que no necesitamos filtrar
                    # Simplemente guardamos los archivos como ProcessedFile
                    ProcessedFile.objects.create(
                        user=request.user,
                        study=study,
                        processing_session=processing_session,
                        original_name=excel_file,
                        file=f'processed_files/{request.user.id}/{study_id}/{processing_session.id}/{excel_file}'
                    )

                # Actualizar el estado del estudio a 'Procesado'
                study.estado = 'Procesado'
                study.save()

                # Mensaje de éxito y redirección con parámetro 'study'
                messages.success(request, 'Paquetes subidos y procesados exitosamente.')
                return redirect(f"{reverse('upload_file')}?study={study_id}")
        else:
            # Si el formulario no es válido, mantener la selección del estudio
            study_id = request.POST.get('study')
    else:
        # Inicializar el formulario vacío o con la selección del estudio si está presente
        if study_id:
            try:
                selected_study = Study.objects.get(id=study_id, user=request.user)
                form = UploadFileForm(initial={'study': selected_study}, user=request.user)
            except Study.DoesNotExist:
                form = UploadFileForm(user=request.user)
                selected_study = None
        else:
            form = UploadFileForm(user=request.user)
            selected_study = None

    if study_id:
        # Obtener las sesiones de procesamiento para el estudio seleccionado
        processing_sessions = ProcessingSession.objects.filter(user=request.user, study_id=study_id).order_by('-processed_at')
        # Obtener el estudio seleccionado
        try:
            selected_study = Study.objects.get(id=study_id, user=request.user)
        except Study.DoesNotExist:
            selected_study = None
    else:
        processing_sessions = None
        selected_study = None

    context = {
        'form': form,
        'studies': studies,
        'processing_sessions': processing_sessions,
        'selected_study': selected_study,
    }

    return render(request, 'upload.html', context)

@login_required
@user_is_approved
def download_zip_session(request, session_id):
    try:
        session = ProcessingSession.objects.get(id=session_id, user=request.user)
    except ProcessingSession.DoesNotExist:
        return HttpResponse("Session not found.", status=404)

    processed_files = session.processed_files.all()

    zip_filename = f"processed_files_session_{session_id}.zip"
    s = BytesIO()

    with ZipFile(s, "w") as zf:
        for processed_file in processed_files:
            file_path = processed_file.file.path
            if os.path.exists(file_path):
                zf.write(file_path, arcname=processed_file.original_name)

    response = HttpResponse(s.getvalue(), content_type="application/x-zip-compressed")
    response['Content-Disposition'] = f'attachment; filename={zip_filename}'

    return response

@login_required
@user_is_approved
def send_emails(request):
    # Obtener estudios ordenados por expediente
    studies = Study.objects.filter(user=request.user).order_by('-expediente')

    # Obtener el estudio seleccionado, ya sea desde POST o GET
    selected_study_id = request.POST.get('study') if request.method == 'POST' else request.GET.get('study')
    selected_study = Study.objects.filter(id=selected_study_id, user=request.user).first() if selected_study_id else None

    already_sent = False
    sent_files = []
    not_processed = False
    recent_files = []
    sent_at = None

    contact_success_rates = {}  # Diccionario para almacenar la tasa de éxito por contacto

    if selected_study:
        # Verificar si ya se han enviado correos para este estudio
        sent_email_records = SentEmail.objects.filter(user=request.user, study=selected_study).order_by('-sent_at')
        if sent_email_records.exists():
            already_sent = True
            latest_sent_email = sent_email_records.first()
            sent_at = latest_sent_email.sent_at

            # Recopilar archivos enviados y destinatarios
            files = ProcessedFile.objects.filter(user=request.user, study=selected_study).order_by('original_name')
            for file in files:
                emails_with_file = SentEmail.objects.filter(user=request.user, study=selected_study, files=file)
                recipients = []
                for email_record in emails_with_file:
                    recipient_emails = email_record.to_emails.split(',')
                    recipient_names = email_record.to_names.split(',') if email_record.to_names else recipient_emails
                    for i, recipient_email in enumerate(recipient_emails):
                        recipient_email = recipient_email.strip()
                        recipient_name = recipient_names[i].strip() if i < len(recipient_names) else recipient_email
                        recipients.append({'name': recipient_name, 'email': recipient_email})
                sent_files.append({'file': file, 'recipients': recipients})

        # Comprobar si el estudio ha sido procesado
        if selected_study.estado == 'Pendiente':
            not_processed = True
        else:
            # Obtener archivos procesados y ordenarlos alfabéticamente
            recent_files_qs = ProcessedFile.objects.filter(user=request.user, study=selected_study).order_by('original_name')
            recent_files_data = []
            for file in recent_files_qs:
                # Obtener los destinatarios y sus estados
                contact_statuses = ProcessedFileContactStatus.objects.filter(processed_file=file).select_related('contact')
                recipients_data = []
                for cs in contact_statuses:
                    recipients_data.append({
                        'id': cs.contact.id,  # Agregado para acceder al ID en la plantilla
                        'name': cs.contact.name,
                        'email': cs.contact.email,
                        'status': cs.status,
                    })

                # Agregar a recent_files_data
                recent_files_data.append({
                    'original_name': file.original_name,
                    'id': file.id,  # Agregado para acceder al ID en la plantilla
                    'recipients': recipients_data,
                })
            recent_files = recent_files_data

    # Obtener todos los contactos del usuario, ordenados alfabéticamente por nombre
    contacts = Contact.objects.filter(user=request.user).order_by('name')

    # Calcular la tasa de éxito para cada contacto
    for contact in contacts:
        total_sent = ProcessedFileContactStatus.objects.filter(contact=contact).count()
        total_received = ProcessedFileContactStatus.objects.filter(contact=contact, status='Recibido').count()
        success_rate = (total_received / total_sent) * 100 if total_sent > 0 else 0
        contact_success_rates[contact.id] = round(success_rate, 2)  # Redondear a 2 decimales

    # Generar colores para cada gremio asegurando que sean claros y distintos
    gremio_colors = {}
    unique_gremios = set()

    for contact in contacts:
        if contact.gremio:
            unique_gremios.update([g.strip() for g in contact.gremio.split(',')])

    unique_gremios = sorted(unique_gremios)

    num_gremios = len(unique_gremios)
    for index, gremio in enumerate(unique_gremios):
        hue = index / num_gremios  # Distribuir los tonos equitativamente
        lightness = 0.8  # Asegurar que los colores sean claros
        saturation = 0.6  # Saturación media para colores vivos
        r, g, b = colorsys.hls_to_rgb(hue, lightness, saturation)
        r_hex = int(r * 255)
        g_hex = int(g * 255)
        b_hex = int(b * 255)
        gremio_colors[gremio] = f'#{r_hex:02x}{g_hex:02x}{b_hex:02x}'

    # Procesar el envío de correos si el formulario es enviado
    if request.method == 'POST' and ('send_email' in request.POST or 'send_and_download' in request.POST):
        form = SendEmailForm(request.POST, user=request.user)
        if form.is_valid():
            selected_study = form.cleaned_data['study']
            if selected_study.estado == 'Pendiente':
                messages.warning(request, "Este estudio aún no ha sido procesado.")
                return redirect('send_emails')
            else:
                # Lógica de envío de correos
                body_content = form.cleaned_data['body']
                clean_body = html.unescape(body_content).replace('<br>', '\n').replace('<br/>', '\n').replace('<br />', '\n')

                try:
                    subject = form.cleaned_data['subject']
                    from_email_account = EmailAccount.objects.get(id=form.cleaned_data['from_email'], user=request.user)
                    cc_emails = [email.strip() for email in form.cleaned_data.get('cc', '').split(',') if email.strip()]
                    bcc_emails = [email.strip() for email in form.cleaned_data.get('bcc', '').split(',') if email.strip()]

                    contacts_files = {}
                    for file in recent_files:
                        recipient_key = f'recipients_{file["id"]}[]'
                        recipient_ids = request.POST.getlist(recipient_key)

                        for recipient_id in recipient_ids:
                            if recipient_id not in contacts_files:
                                contacts_files[recipient_id] = []
                            contacts_files[recipient_id].append(file["id"])

                    zip_buffer = BytesIO()
                    with zipfile.ZipFile(zip_buffer, 'w') as zf:
                        for recipient_id, file_ids in contacts_files.items():
                            contact = Contact.objects.get(id=recipient_id)
                            to_email = contact.email
                            to_name = contact.name

                            email_message = EmailMultiAlternatives(
                                subject=subject,
                                body=clean_body,
                                from_email=from_email_account.email,
                                to=[to_email],
                                cc=cc_emails,
                                bcc=bcc_emails,
                                headers={'Reply-To': from_email_account.email},
                            )

                            email_message.attach_alternative(body_content, "text/html")

                            # Adjuntar los archivos seleccionados
                            for file_id in file_ids:
                                file = ProcessedFile.objects.get(id=file_id)
                                email_message.attach_file(file.file.path)

                            connection = get_connection(
                                backend='django.core.mail.backends.smtp.EmailBackend',
                                host=from_email_account.smtp_server,
                                port=from_email_account.smtp_port,
                                username=from_email_account.smtp_username,
                                password=from_email_account.smtp_password,
                                use_tls=True
                            )

                            email_message.connection = connection
                            email_message.send()

                            # Crear registro de correo enviado
                            sent_email = SentEmail.objects.create(
                                user=request.user,
                                from_email=from_email_account.email,
                                to_emails=to_email,
                                to_names=to_name,  # Guardar el nombre del contacto
                                subject=subject,
                                body=body_content,
                                study=selected_study
                            )

                            for file_id in file_ids:
                                file = ProcessedFile.objects.get(id=file_id)
                                sent_email.files.add(file)

                            # Crear o actualizar ProcessedFileContactStatus
                            for file_id in file_ids:
                                file = ProcessedFile.objects.get(id=file_id)
                                ProcessedFileContactStatus.objects.update_or_create(
                                    processed_file=file,
                                    contact=contact,
                                    defaults={'status': 'Pendiente'}
                                )

                    # Actualizar el estado del estudio a 'Enviado'
                    selected_study.estado = 'Enviado'
                    selected_study.save()

                    if 'send_and_download' in request.POST:
                        zip_buffer.seek(0)
                        response = HttpResponse(zip_buffer, content_type='application/zip')
                        response['Content-Disposition'] = 'attachment; filename=correos_enviados.zip'
                        messages.success(request, 'Correos enviados exitosamente. Se ha generado un archivo ZIP para descargar.')
                        return response
                    else:
                        messages.success(request, 'Correos enviados exitosamente.')
                        return redirect('send_emails')

                except Exception as e:
                    messages.error(request, f'Error al enviar el correo: {e}')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = SendEmailForm(user=request.user)
        if selected_study_id:
            form.fields['study'].initial = selected_study_id
        email_accounts = EmailAccount.objects.filter(user=request.user)
        if email_accounts.exists():
            form.fields['from_email'].initial = email_accounts.first().id

    # Crear un diccionario para mantener las selecciones de contactos por archivo
    selected_contacts = {}
    if request.method == 'POST':
        for key in request.POST:
            if key.startswith('recipients_'):
                file_id = key.split('_')[1].rstrip('[]')
                selected_contacts.setdefault(file_id, set()).update(request.POST.getlist(key))

    if not contacts.exists() and selected_study:
        messages.warning(request, "No se encontraron contactos. Por favor, añade algunos antes de enviar correos.")

    context = {
        'form': form,
        'studies': studies,
        'selected_study_id': selected_study_id,
        'selected_study': selected_study,
        'not_processed': not_processed,
        'already_sent': already_sent,
        'sent_at': sent_at,
        'sent_files': sent_files,
        'recent_files': recent_files,
        'contacts': contacts,
        'gremio_colors': gremio_colors,
        'contact_success_rates': contact_success_rates,  # Añadido para pasar las tasas de éxito
        'selected_contacts': selected_contacts,  # Añadido para mantener las selecciones
    }

    return render(request, 'send_emails.html', context)

@login_required
@user_is_approved
def download_emails_zip(request):
    # Crea un buffer en memoria para el archivo zip
    zip_buffer = BytesIO()

    # Crear el archivo zip en memoria
    with zipfile.ZipFile(zip_buffer, 'w') as zf:
        # Obtén todos los correos enviados por el usuario actual
        sent_emails = SentEmail.objects.filter(user=request.user)
        
        for email in sent_emails:
            # Crear el contenido del archivo .eml
            eml_content = f"From: {email.from_email}\nTo: {', '.join(email.to_emails.split(','))}\nSubject: {email.subject}\n\n{email.body}"
            
            # Definir un nombre único para cada correo en el zip
            eml_filename = f"{email.subject[:50]}_{email.id}.eml"
            
            # Agregar el archivo al zip
            zf.writestr(eml_filename, eml_content)

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=correos_enviados.zip'
    return response

@login_required
@user_is_approved
def contact_list(request):
    # Obtener todos los contactos del usuario actual
    contacts = Contact.objects.filter(user=request.user)

    # Crear una lista para almacenar contactos con sus gremios y colores
    contacts_with_gremios = []

    # Obtener todos los gremios únicos
    unique_gremios = set()
    for contact in contacts:
        if contact.gremio:
            gremio_list = [g.strip() for g in contact.gremio.split(',')]
            unique_gremios.update(gremio_list)

    unique_gremios = sorted(unique_gremios)
    num_gremios = len(unique_gremios) if unique_gremios else 1  # Evitar división por cero

    # Asignar colores basados en hue para que sean diferentes y distribuidos como arco iris
    gremio_colors = {}
    for index, gremio in enumerate(unique_gremios):
        hue = (index / num_gremios)  # Distribuir los tonos equitativamente (0-1)
        lightness = 0.8  # Ajustar para asegurar que los colores sean claros
        saturation = 0.6  # Saturación media para colores vivos
        r, g, b = colorsys.hls_to_rgb(hue, lightness, saturation)
        r_hex = int(r * 255)
        g_hex = int(g * 255)
        b_hex = int(b * 255)
        gremio_colors[gremio] = f'#{r_hex:02x}{g_hex:02x}{b_hex:02x}'

    # Construir la lista de contactos con sus gremios y colores
    for contact in contacts:
        gremios = [g.strip() for g in contact.gremio.split(',')] if contact.gremio else []
        colored_gremios = []
        for gremio in gremios:
            color = gremio_colors.get(gremio, '#d3d3d3')  # Color gris por defecto
            colored_gremios.append({
                'name': gremio,
                'background_color': color,
                'text_color': '#000000',  # Texto siempre negro
            })
        contacts_with_gremios.append({'contact': contact, 'gremios': colored_gremios})

    # Variables para manejo de edición
    edit_contact = None
    if 'edit_id' in request.GET:
        contact_id = request.GET.get('edit_id')
        edit_contact = get_object_or_404(Contact, id=contact_id, user=request.user)

    if request.method == 'POST':
        if 'delete_contact' in request.POST:
            # Obtener el id del contacto a eliminar
            contact_id = request.POST.get('contact_id')
            contact = get_object_or_404(Contact, id=contact_id, user=request.user)
            contact.delete()
            messages.success(request, 'Contacto eliminado exitosamente.')
            return redirect('contact_list')
        elif 'add_contact' in request.POST:
            # Manejar la creación de un nuevo contacto
            name = request.POST.get('name')
            email = request.POST.get('email')
            phone = request.POST.get('phone')
            gremio = request.POST.get('gremio')
            if not name or not email:
                messages.error(request, 'Nombre y Email son obligatorios.')
                return redirect('contact_list')
            Contact.objects.create(name=name, email=email, phone=phone, gremio=gremio, user=request.user)
            messages.success(request, 'Contacto agregado exitosamente.')
            return redirect('contact_list')
        elif 'save_changes' in request.POST:
            # Manejar la edición de un contacto existente
            contact_id = request.POST.get('edit_id')
            contact = get_object_or_404(Contact, id=contact_id, user=request.user)
            name = request.POST.get('name')
            email = request.POST.get('email')
            phone = request.POST.get('phone')
            gremio = request.POST.get('gremio')
            if not name or not email:
                messages.error(request, 'Nombre y Email son obligatorios.')
                return redirect('contact_list')
            contact.name = name
            contact.email = email
            contact.phone = phone
            contact.gremio = gremio
            contact.save()
            messages.success(request, 'Contacto actualizado exitosamente.')
            return redirect('contact_list')

    context = {
        'contacts_with_gremios': contacts_with_gremios,
        'gremio_colors': gremio_colors,  # Pasar los colores al contexto
        'edit_contact': edit_contact,    # Pasar el contacto a editar al contexto
    }

    return render(request, 'contact_list.html', context)

@login_required
@user_is_approved
def calendar(request):
    return render(request, 'calendar.html')

@login_required
@user_is_approved
def studies(request):
    study_to_edit = None  # Variable para almacenar el estudio a editar
    edit_mode = False  # Indica si estamos en modo edición

    # Manejar la sumisión del formulario
    if request.method == 'POST':
        # Verificar si el usuario quiere editar un estudio
        if 'edit-study' in request.POST:
            study_id = request.POST.get('study-id')
            study = get_object_or_404(Study, id=study_id, user=request.user)
            form = StudyForm(request.POST, instance=study)
            edit_mode = True  # Cambiar al modo edición
            study_to_edit = study  # Almacenar el estudio a editar
            if form.is_valid():
                form.save()
                messages.success(request, 'Estudio actualizado exitosamente.')
                return redirect('studies')

        # Verificar si el usuario quiere eliminar un estudio
        elif 'delete-study' in request.POST:
            study_id = request.POST.get('study-id')
            study = get_object_or_404(Study, id=study_id, user=request.user)
            study.delete()
            messages.success(request, 'Estudio eliminado exitosamente.')
            return redirect('studies')
        
        # Manejar el envío de un estudio
        elif 'send-study' in request.POST:
            study_id = request.POST.get('study-id')
            study = get_object_or_404(Study, id=study_id, user=request.user)
            # Redirigir a la vista de enviar correos
            return redirect('send_emails', study_id=study.id)

        # Manejar la reclamación de un estudio
        elif 'reclaim-study' in request.POST:
            study_id = request.POST.get('study-id')
            study = get_object_or_404(Study, id=study_id, user=request.user)
            # Redirigir a la vista de reclamar estudio
            return redirect('reclaim_study', study_id=study.id)

        # Manejar agregar un nuevo estudio
        else:
            form = StudyForm(request.POST)
            if form.is_valid():
                study = form.save(commit=False)
                study.user = request.user
                study.save()
                messages.success(request, 'Estudio agregado exitosamente.')
                return redirect('studies')

    else:
        form = StudyForm()  # Formulario vacío por defecto

    # Filtros de búsqueda y ordenamiento
    query = request.GET.get('query', '')
    sort_by = request.GET.get('sort_by', 'expediente')
    order = request.GET.get('order', 'asc')

    # Definir órdenes de columna para alternar entre ascendente y descendente
    expediente_order = 'asc' if sort_by == 'expediente' and order == 'desc' else 'desc'
    project_name_order = 'asc' if sort_by == 'project_name' and order == 'desc' else 'desc'
    due_date_order = 'asc' if sort_by == 'due_date' and order == 'desc' else 'desc'
    estado_order = 'asc' if sort_by == 'estado' and order == 'desc' else 'desc'

    if order == 'desc':
        sort_by = f'-{sort_by}'

    # Filtrado y ordenamiento
    studies = Study.objects.filter(
        Q(user=request.user),
        Q(expediente__icontains=query) |
        Q(project_name__icontains=query)
    ).order_by(sort_by)

    # Obtener el estudio a editar si está en modo edición
    if edit_mode and study_to_edit:
        # El formulario ya está inicializado con el estudio a editar
        pass
    else:
        study_to_edit = None

    # Contexto para la plantilla
    context = {
        'form': form,
        'studies': studies,
        'study_to_edit': study_to_edit,
        'edit_mode': edit_mode,
        'expediente_order': expediente_order,
        'project_name_order': project_name_order,
        'due_date_order': due_date_order,
        'estado_order': estado_order,
        'current_sort': sort_by.lstrip('-'),
        'current_order': order,
        'query': query,
    }

    return render(request, 'studies.html', context)

@login_required
@user_is_approved
def preview_file(request, file_id):
    processed_file = get_object_or_404(ProcessedFile, id=file_id, user=request.user)
    file_path = processed_file.file.path

    # Cargar el archivo Excel
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    # Convertir el contenido del archivo Excel en una lista de listas
    data = []
    for row in ws.iter_rows(values_only=True):
        filtered_row = [cell if cell is not None else '' for cell in row]
        if any(filtered_row):  # Asegurarse de no agregar filas completamente vacías
            data.append(filtered_row)

    context = {
        'file_name': processed_file.original_name,
        'data': data
    }

    return render(request, 'preview.html', context)

@login_required
@user_is_approved
def add_email_account(request):
    if request.method == 'POST':
        form = EmailAccountForm(request.POST)
        if form.is_valid():
            email_account = form.save(commit=False)
            email_account.user = request.user
            email_account.save()
            return redirect('send_emails')  # Redirect to send emails page after saving
    else:
        form = EmailAccountForm()

    return render(request, 'add_email_account.html', {'form': form})

@login_required
@user_is_approved
def edit_email_account(request, account_id):
    email_account = get_object_or_404(EmailAccount, id=account_id, user=request.user)
    
    if request.method == 'POST':
        form = EmailAccountForm(request.POST, instance=email_account)
        if form.is_valid():
            form.save()
            messages.success(request, 'Credenciales de correo actualizadas con éxito.')
            return redirect('list_email_accounts')  # Redirigir a la lista de cuentas de correo o cualquier vista adecuada
        else:
            messages.error(request, 'Hubo un error al actualizar las credenciales. Por favor, corrige los errores en el formulario.')
    else:
        form = EmailAccountForm(instance=email_account)

    return render(request, 'edit_email_account.html', {'form': form})

@login_required
@user_is_approved
def list_email_accounts(request):
    email_accounts = EmailAccount.objects.filter(user=request.user)
    return render(request, 'list_email_accounts.html', {'email_accounts': email_accounts})

@login_required
@user_is_approved
def study_details(request, study_id):
    # Obtener el estudio
    study = get_object_or_404(Study, id=study_id, user=request.user)
    
    # Obtener los archivos procesados del estudio ordenados alfabéticamente
    processed_files = ProcessedFile.objects.filter(study=study).order_by('original_name')
    
    # Crear un diccionario para almacenar los correos enviados y los destinatarios de cada archivo
    correos_enviados = {}
    
    # Recorrer cada archivo procesado y obtener los destinatarios reales
    for file in processed_files:
        # Obtener todos los ProcessedFileContactStatus para este archivo
        contact_statuses = ProcessedFileContactStatus.objects.filter(processed_file=file).select_related('contact')
        
        # Solo agregar el archivo si hay contactos asociados
        if contact_statuses.exists():
            correos_enviados[file.id] = {
                'file': file,
                'recipients': [],
                'file_status': 'Pendiente'
            }
            
            # Recopilar información de cada contacto
            for cs in contact_statuses:
                correos_enviados[file.id]['recipients'].append({
                    'id': cs.id,  # ID de ProcessedFileContactStatus
                    'name': cs.contact.name,
                    'email': cs.contact.email,
                    'phone': cs.contact.phone,
                    'status': cs.status,
                })
            
            # Determinar el estado global del archivo basado en los destinatarios
            recipient_count = len(correos_enviados[file.id]['recipients'])
            reclamados = sum(1 for r in correos_enviados[file.id]['recipients'] if r['status'] == 'Reclamado')
            recibidos = sum(1 for r in correos_enviados[file.id]['recipients'] if r['status'] == 'Recibido')
            rechazados = sum(1 for r in correos_enviados[file.id]['recipients'] if r['status'] == 'Rechazado')
            
            if reclamados == recipient_count and recipient_count > 0:
                correos_enviados[file.id]['file_status'] = 'Reclamado'
            elif recibidos == recipient_count and recipient_count > 0:
                correos_enviados[file.id]['file_status'] = 'Recibido'
            elif rechazados == recipient_count and recipient_count > 0:
                correos_enviados[file.id]['file_status'] = 'Rechazado'
            else:
                correos_enviados[file.id]['file_status'] = 'Pendiente'
    
    if request.method == 'POST':
        # Manejar solicitudes AJAX para actualizar el estado de un destinatario
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            recipient_id = request.POST.get('recipient_id')
            file_id = request.POST.get('file_id')
            nuevo_estado = request.POST.get('nuevo_estado')
    
            if recipient_id and file_id and nuevo_estado:
                # Validar que el nuevo estado es válido
                valid_statuses = [choice[0] for choice in ProcessedFileContactStatus.STATUS_CHOICES]
                if nuevo_estado not in valid_statuses:
                    return JsonResponse({'error': 'Estado inválido.'}, status=400)
    
                try:
                    # Convertir IDs a enteros
                    file_id = int(file_id)
                    recipient_id = int(recipient_id)
                except ValueError:
                    return JsonResponse({'error': 'IDs inválidos.'}, status=400)
    
                try:
                    # Obtener el registro específico de ProcessedFileContactStatus
                    contact_status = ProcessedFileContactStatus.objects.get(
                        id=recipient_id, 
                        processed_file_id=file_id, 
                        processed_file__study__user=request.user
                    )
                    # Actualizar el estado
                    contact_status.status = nuevo_estado
                    contact_status.save()
                    
                    # Recalcular el estado global del archivo después de la actualización
                    file = contact_status.processed_file
                    contact_statuses = ProcessedFileContactStatus.objects.filter(processed_file=file)
                    statuses = contact_statuses.values_list('status', flat=True)
                    unique_statuses = set(statuses)
                    
                    if len(unique_statuses) == 1 and 'Pendiente' not in unique_statuses:
                        file_status = unique_statuses.pop()
                    else:
                        file_status = 'Pendiente'
                    
                    return JsonResponse({'message': 'Estado actualizado correctamente.', 'file_status': file_status})
                except ProcessedFileContactStatus.DoesNotExist:
                    return JsonResponse({'error': 'Estado no encontrado.'}, status=404)
                except Exception as e:
                    # Para depuración, puedes incluir el mensaje de excepción
                    return JsonResponse({'error': f'Ocurrió un error inesperado: {str(e)}'}, status=500)
            return JsonResponse({'error': 'Datos inválidos.'}, status=400)
    
    # Renderizar la plantilla con los datos procesados
    context = {
        'study': study,
        'processed_files': processed_files,
        'correos_enviados': correos_enviados,
        'estado': study.estado
    }
    return render(request, 'study_details.html', context)

@require_POST
@login_required
@user_is_approved
def update_estado_contacto(request):
    recipient_id = request.POST.get('recipient_id')
    file_id = request.POST.get('file_id')
    nuevo_estado = request.POST.get('nuevo_estado')

    if recipient_id and file_id and nuevo_estado:
        # Validar que el nuevo estado es válido
        valid_statuses = [choice[0] for choice in ProcessedFileContactStatus.STATUS_CHOICES]
        if nuevo_estado not in valid_statuses:
            return JsonResponse({'error': 'Estado inválido.'}, status=400)

        try:
            # Convertir IDs a enteros
            file_id = int(file_id)
            recipient_id = int(recipient_id)
        except ValueError:
            return JsonResponse({'error': 'IDs inválidos.'}, status=400)

        try:
            # Obtener el registro específico de ProcessedFileContactStatus
            contact_status = ProcessedFileContactStatus.objects.get(
                id=recipient_id, 
                processed_file_id=file_id, 
                processed_file__study__user=request.user
            )
            # Actualizar el estado
            contact_status.status = nuevo_estado
            contact_status.save()
            
            # Recalcular el estado global del archivo después de la actualización
            file = contact_status.processed_file
            contact_statuses = ProcessedFileContactStatus.objects.filter(processed_file=file)
            statuses = contact_statuses.values_list('status', flat=True)
            unique_statuses = set(statuses)
            
            if len(unique_statuses) == 1 and 'Pendiente' not in unique_statuses:
                file_status = unique_statuses.pop()
            else:
                file_status = 'Pendiente'
            
            return JsonResponse({'message': 'Estado actualizado correctamente.', 'file_status': file_status})
        except ProcessedFileContactStatus.DoesNotExist:
            return JsonResponse({'error': 'Estado no encontrado.'}, status=404)
        except Exception as e:
            # Para depuración, puedes incluir el mensaje de excepción
            return JsonResponse({'error': f'Ocurrió un error inesperado: {str(e)}'}, status=500)
    return JsonResponse({'error': 'Datos inválidos.'}, status=400)

@login_required
@user_is_approved
def delete_contact(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id, user=request.user)
    contact.delete()
    return redirect('contact_list')

logger = logging.getLogger(__name__)

@login_required
@user_is_approved
def calendar_view(request):
    # Get all studies and events
    studies = Study.objects.all()
    events = Event.objects.filter(user=request.user)

    # Create a dictionary to store study colors
    study_colors = {}

    # Define a function to generate a light color for each study
    def generate_color():
        return "#{:02x}{:02x}{:02x}".format(random.randint(180, 255), random.randint(180, 255), random.randint(180, 255))  # Light colors

    # Assign colors to studies
    for study in studies:
        if study.id not in study_colors:
            study_colors[study.id] = generate_color()

    # Collect events for the calendar
    event_list = []

    # For delivery events (ENTREGA)
    for study in studies:
        if study.due_date:
            event_list.append({
                'id': f'entrega-{study.id}',
                'title': f'{study.expediente} - ENTREGA',  # Correct format for title
                'start': study.due_date.strftime('%Y-%m-%d %H:%M'),
                'backgroundColor': study_colors[study.id],  # Use study color for background
                'borderColor': study_colors[study.id],  # Optional: Make border same as background
                'textColor': '#000',  # Set text color to black
                'font-weight': 'bold',
                'study_id': study.id,
                'is_delivery': True  # Mark this event as a delivery
            })
            
    # For regular events related to a study
    for event in events:
        if event.study:
            event_list.append({
                'id': event.id,
                'title': f'{event.study.expediente} - {event.title}',  # Keep expediente and title
                'start': event.start_date.strftime('%Y-%m-%d %H:%M'),
                'end': event.end_date.strftime('%Y-%m-%d %H:%M') if event.end_date else None,
                'color': study_colors[event.study.id],  # Match color with the study
                'backgroundColor': study_colors[event.study.id],  # Ensure color is applied
                'textColor': '#000',  # Black text for readability
                'study_id': event.study.id,
                'expediente': event.study.expediente,
                'is_delivery': False
            })
        else:
            # For events without assigned study, use a distinct color
            event_list.append({
                'id': event.id,
                'title': f'{event.title}',  # Just event title
                'start': event.start_date.strftime('%Y-%m-%d %H:%M'),
                'end': event.end_date.strftime('%Y-%m-%d %H:%M') if event.end_date else None,
                'color': '#d3d3d3',  # Light gray for unassigned events
                'backgroundColor': '#d3d3d3',  # Ensure color is applied
                'textColor': '#000',  # Black text for readability
                'study_id': None,
                'is_delivery': False
            })

    form = EventForm()

    context = {
        'form': form,
        'events': json.dumps(event_list),
        'studies': studies,
    }

    return render(request, 'calendar.html', context)

@login_required
@user_is_approved
def calendar_day_view(request, date):
    # Parse the date from the URL parameter
    date_obj = parse_date(date)
    
    if not date_obj:
        return render(request, '404.html', status=404)
    
    # Get all events for the selected day
    events = Event.objects.filter(start_date__date=date_obj)
    
    context = {
        'events': events,
        'date': date_obj,
    }
    
    return render(request, 'calendar_day.html', context)

logger = logging.getLogger(__name__)

@login_required
@user_is_approved
def reclamar(request):
    # Obtener estudios ordenados por expediente
    studies = Study.objects.filter(user=request.user).order_by('-expediente')

    # Obtener el estudio seleccionado, ya sea desde POST o GET
    selected_study_id = request.POST.get('study') if request.method == 'POST' else request.GET.get('study')
    selected_study = Study.objects.filter(id=selected_study_id, user=request.user).first() if selected_study_id else None

    already_sent = False
    sent_files = []
    sent_at = None
    already_claimed = False
    claimed_files = []
    claimed_at = None
    study_not_processed = False
    recent_files = []

    contact_success_rates = {}  # Diccionario para almacenar la tasa de éxito por contacto

    if selected_study:
        # Verificar si ya se han enviado correos para este estudio (Envío)
        sent_email_records = SentEmail.objects.filter(user=request.user, study=selected_study, email_type='Envío').order_by('-sent_at')
        if sent_email_records.exists():
            already_sent = True
            latest_sent_email = sent_email_records.first()
            sent_at = latest_sent_email.sent_at

            # Recopilar archivos enviados y destinatarios
            files = ProcessedFile.objects.filter(user=request.user, study=selected_study).order_by('original_name')
            for file in files:
                emails_with_file = SentEmail.objects.filter(user=request.user, study=selected_study, files=file, email_type='Envío')
                recipients = []
                for email_record in emails_with_file:
                    recipient_emails = email_record.to_emails.split(',')
                    recipient_names = email_record.to_names.split(',') if email_record.to_names else recipient_emails
                    for i, recipient_email in enumerate(recipient_emails):
                        recipient_email = recipient_email.strip()
                        recipient_name = recipient_names[i].strip() if i < len(recipient_names) else recipient_email
                        recipients.append({'name': recipient_name, 'email': recipient_email})
                sent_files.append({'file': file, 'recipients': recipients})

        # Verificar si ya se han enviado reclamaciones para este estudio (Reclamo)
        claimed_email_records = SentEmail.objects.filter(user=request.user, study=selected_study, email_type='Reclamo').order_by('-sent_at')
        if claimed_email_records.exists():
            already_claimed = True
            latest_claimed_email = claimed_email_records.first()
            claimed_at = latest_claimed_email.sent_at

            # Recopilar archivos reclamados y destinatarios
            files = ProcessedFile.objects.filter(user=request.user, study=selected_study).order_by('original_name')
            for file in files:
                # Obtener contactos reclamados para este archivo
                reclaimed_statuses = ProcessedFileContactStatus.objects.filter(
                    processed_file=file,
                    status='Reclamado'
                ).select_related('contact')
                recipients = []
                for status in reclaimed_statuses:
                    recipient = status.contact
                    recipients.append({'id': recipient.id, 'name': recipient.name, 'email': recipient.email})
                claimed_files.append({'file': file, 'recipients': recipients})

        # Comprobar si el estudio ha sido procesado y enviado
        if selected_study.estado == 'Pendiente':
            study_not_processed = True
        else:
            # Obtener archivos procesados y ordenarlos alfabéticamente
            processed_files = ProcessedFile.objects.filter(user=request.user, study=selected_study).order_by('original_name')
            recent_files_data = []
            for file in processed_files:
                # Obtener los destinatarios y sus estados
                contact_statuses = ProcessedFileContactStatus.objects.filter(
                    processed_file=file,
                    status__in=['Pendiente', 'Enviado', 'Reclamado']
                ).select_related('contact')
                if not contact_statuses.exists():
                    continue  # Saltar este archivo si no tiene contactos en los estados deseados
                recipients_data = []
                recipient_ids = []
                reclaimed_contact_ids = set()  # Lista para almacenar IDs de contactos reclamados
                for cs in contact_statuses:
                    recipients_data.append({
                        'id': cs.contact.id,
                        'name': cs.contact.name,
                        'email': cs.contact.email,
                        'status': cs.status,
                        'gremio': cs.contact.gremio  # Añadir gremio si es necesario en la plantilla
                    })
                    recipient_ids.append(cs.contact.id)
                    if cs.status == 'Reclamado':
                        reclaimed_contact_ids.add(cs.contact.id)

                # Agregar a recent_files_data
                recent_files_data.append({
                    'original_name': file.original_name,
                    'id': file.id,
                    'recipients': recipients_data,
                    'recipient_ids': recipient_ids,
                    'reclaimed_contact_ids': reclaimed_contact_ids,
                })
            recent_files = recent_files_data

    # Obtener contactos asociados al estudio y con estado 'Pendiente', 'Enviado' o 'Reclamado'
    if selected_study:
        contacts = Contact.objects.filter(
            id__in=ProcessedFileContactStatus.objects.filter(
                processed_file__study=selected_study,
                status__in=['Pendiente', 'Enviado', 'Reclamado']
            ).values_list('contact_id', flat=True)
        ).order_by('name')
    else:
        contacts = Contact.objects.none()

    # Calcular la tasa de éxito para cada contacto
    for contact in contacts:
        total_sent = ProcessedFileContactStatus.objects.filter(contact=contact).count()
        total_received = ProcessedFileContactStatus.objects.filter(contact=contact, status='Recibido').count()
        success_rate = (total_received / total_sent) * 100 if total_sent > 0 else 0
        contact_success_rates[contact.id] = round(success_rate, 2)  # Redondear a 2 decimales

    # Generar colores para cada gremio asegurando que sean claros y distintos
    gremio_colors = {}
    unique_gremios = set()

    for contact in contacts:
        if contact.gremio:
            unique_gremios.update([g.strip() for g in contact.gremio.split(',')])

    unique_gremios = sorted(unique_gremios)

    num_gremios = len(unique_gremios) if unique_gremios else 1
    for index, gremio in enumerate(unique_gremios):
        hue = index / num_gremios  # Distribuir los tonos equitativamente
        lightness = 0.8  # Asegurar que los colores sean claros
        saturation = 0.6  # Saturación media para colores vivos
        r, g, b = colorsys.hls_to_rgb(hue, lightness, saturation)
        r_hex = int(r * 255)
        g_hex = int(g * 255)
        b_hex = int(b * 255)
        gremio_colors[gremio] = f'#{r_hex:02x}{g_hex:02x}{b_hex:02x}'

    # Procesar el envío de correos de reclamación si el formulario es enviado
    if request.method == 'POST':
        print('Recibida solicitud POST')
        print('Datos POST:', request.POST)
        form = SendEmailForm(request.POST, user=request.user)
        if form.is_valid():
            selected_study = form.cleaned_data['study']
            if selected_study.estado == 'Pendiente':
                messages.warning(request, "Este estudio aún no ha sido enviado. No se pueden enviar reclamaciones.")
                return redirect('reclamar')
            else:
                # Lógica de envío de correos de reclamación
                body_content = form.cleaned_data['body']
                clean_body = html.unescape(body_content).replace('<br>', '\n').replace('<br/>', '\n').replace('<br />', '\n')

                try:
                    subject = form.cleaned_data['subject']
                    from_email_account = form.cleaned_data['from_email']  # Usar directamente el objeto EmailAccount
                    cc_emails = [email.strip() for email in form.cleaned_data.get('cc', '').split(',') if email.strip()]
                    bcc_emails = [email.strip() for email in form.cleaned_data.get('bcc', '').split(',') if email.strip()]

                    contacts_files = {}
                    for file in recent_files:
                        recipient_key = f'recipients_{file["id"]}[]'
                        recipient_ids = request.POST.getlist(recipient_key)

                        for recipient_id in recipient_ids:
                            if recipient_id not in contacts_files:
                                contacts_files[recipient_id] = []
                            contacts_files[recipient_id].append(file["id"])

                    for recipient_id, file_ids in contacts_files.items():
                        contact = Contact.objects.get(id=recipient_id)
                        to_email = contact.email
                        to_name = contact.name

                        email_message = EmailMultiAlternatives(
                            subject=subject,
                            body=clean_body,
                            from_email=from_email_account.email,
                            to=[to_email],
                            cc=cc_emails,
                            bcc=bcc_emails,
                            headers={'Reply-To': from_email_account.email},
                        )

                        email_message.attach_alternative(body_content, "text/html")

                        # Adjuntar los archivos seleccionados
                        for file_id in file_ids:
                            file_obj = ProcessedFile.objects.get(id=file_id)
                            email_message.attach_file(file_obj.file.path)

                        connection = get_connection(
                            backend='django.core.mail.backends.smtp.EmailBackend',
                            host=from_email_account.smtp_server,
                            port=from_email_account.smtp_port,
                            username=from_email_account.smtp_username,
                            password=from_email_account.smtp_password,
                            use_tls=True
                        )

                        email_message.connection = connection
                        email_message.send()

                        # Crear registro de correo enviado
                        sent_email = SentEmail.objects.create(
                            user=request.user,
                            from_email=from_email_account.email,
                            to_emails=to_email,
                            to_names=to_name,  # Guardar el nombre del contacto
                            subject=subject,
                            body=body_content,
                            study=selected_study,
                            email_type='Reclamo'
                        )

                        for file_id in file_ids:
                            file_obj = ProcessedFile.objects.get(id=file_id)
                            sent_email.files.add(file_obj)

                        # Actualizar el estado a 'Reclamado' en ProcessedFileContactStatus
                        for file_id in file_ids:
                            file_obj = ProcessedFile.objects.get(id=file_id)
                            ProcessedFileContactStatus.objects.filter(
                                processed_file=file_obj,
                                contact=contact
                            ).update(status='Reclamado')

                    messages.success(request, 'Reclamos enviados exitosamente.')
                    return redirect('reclamar')

                except Exception as e:
                    logger.error(f"Error al enviar el correo de reclamación: {e}", exc_info=True)
                    messages.error(request, f'Error al enviar el correo de reclamación: {e}')
        else:
            print('Errores del formulario:', form.errors)
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = SendEmailForm(user=request.user)
        if selected_study_id:
            form.fields['study'].initial = selected_study_id
        email_accounts = EmailAccount.objects.filter(user=request.user)
        if email_accounts.exists():
            form.fields['from_email'].initial = email_accounts.first()

    # Crear un diccionario para mantener las selecciones de contactos por archivo
    selected_contacts = {}
    if request.method == 'POST':
        for key in request.POST:
            if key.startswith('recipients_'):
                file_id = key.split('_')[1].rstrip('[]')
                selected_contacts.setdefault(file_id, set()).update(request.POST.getlist(key))

    if selected_study and not contacts.exists():
        messages.warning(request, "No se encontraron contactos asociados al estudio con estados Pendiente, Enviado o Reclamado.")

    context = {
        'form': form,
        'studies': studies,
        'selected_study_id': selected_study_id,
        'selected_study': selected_study,
        'study_not_processed': study_not_processed,
        'already_sent': already_sent,
        'sent_at': sent_at,
        'sent_files': sent_files,
        'already_claimed': already_claimed,
        'claimed_at': claimed_at,
        'claimed_files': claimed_files,
        'recent_files': recent_files,
        'contacts': contacts,
        'gremio_colors': gremio_colors,
        'contact_success_rates': contact_success_rates,
        'selected_contacts': selected_contacts,
    }

    return render(request, 'reclamar.html', context)


@login_required
@user_is_approved
def delete_session(request, session_id):
    session = get_object_or_404(ProcessingSession, id=session_id, user=request.user)
    study_id = session.study.id
    if request.method == 'POST':
        # Delete associated files
        processed_files = session.processed_files.all()
        for file in processed_files:
            # Delete the file from the storage
            if file.file and os.path.isfile(file.file.path):
                os.remove(file.file.path)
            file.delete()
        # Delete the session
        session.delete()
        messages.success(request, 'Sesión eliminada exitosamente.')
        return redirect(f'{reverse("upload_file")}?study={study_id}')
    else:
        return HttpResponse('Método no permitido', status=405)

@login_required
@user_is_approved
def download_file(request, file_id):
    # Obtener el archivo procesado asegurando que pertenece al usuario
    processed_file = get_object_or_404(ProcessedFile, id=file_id, user=request.user)
    file_path = processed_file.file.path

    if os.path.exists(file_path):
        # Preparar la respuesta para descargar el archivo
        response = FileResponse(open(file_path, 'rb'), as_attachment=True, filename=processed_file.original_name)
        return response
    else:
        # Retornar un error 404 si el archivo no existe
        return HttpResponse("Archivo no encontrado.", status=404)

@login_required
@user_is_approved
def export_study_excel(request, study_id):
    try:
        # Obtener el estudio asegurando que pertenece al usuario
        study = get_object_or_404(Study, id=study_id, user=request.user)
        
        # Crear un libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Estudio_{study.expediente}"
        
        # Definir los encabezados de la columna A y sus valores en la columna B
        encabezados = ["Expediente", "Nombre del proyecto:", "Fecha de entrega", "Estado del Estudio"]
        valores = [
            study.expediente,
            study.project_name,
            study.due_date.strftime('%Y-%m-%d') if study.due_date else '',
            study.estado
        ]
        
        # Establecer la fuente a Calibri
        calibri_font = Font(name='Calibri')
        
        # Definir alineaciones
        left_alignment = Alignment(horizontal='left', vertical='center')
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        for row_num, (header, valor) in enumerate(zip(encabezados, valores), start=1):
            cell_a = ws.cell(row=row_num, column=1, value=header)
            cell_b = ws.cell(row=row_num, column=2, value=valor)
            
            # Aplicar fuente Calibri
            cell_a.font = Font(name='Calibri', bold=True)
            cell_b.font = Font(name='Calibri')
            
            # Alinear horizontalmente a la izquierda en la columna A y al centro en la columna B
            cell_a.alignment = left_alignment
            cell_b.alignment = center_alignment
        
        # Dejar una fila en blanco
        current_row = len(encabezados) + 1  # Fila 5
        current_row += 1  # Fila 6
        
        # Escribir los encabezados de la tabla de archivos
        table_headers = ['Nombre del Archivo', 'Nombre del Contacto', 'Correo Electrónico', 'Teléfono', 'Estado']
        for col_num, header in enumerate(table_headers, start=1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.font = Font(name='Calibri', bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # Aplicar color de fondo gris claro para los encabezados
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            # Aplicar bordes a los encabezados
            thin_border = Border(
                left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000')
            )
            cell.border = thin_border
        
        current_row += 1  # Fila 7
        
        # Definir los colores de fondo basados en el estado
        status_colors = {
            'Reclamado': 'FFF3CD',  # Light Yellow
            'Recibido': 'D4EDDA',    # Light Green
            'Rechazado': 'F8D7DA',   # Light Red
            'Pendiente': 'F2F2F2'    # Light Gray
        }
        
        # Obtener los datos de los archivos y sus destinatarios
        processed_files = ProcessedFile.objects.filter(study=study).order_by('original_name')
        
        for file in processed_files:
            contact_statuses = ProcessedFileContactStatus.objects.filter(processed_file=file).select_related('contact').order_by('id')
            recipients = contact_statuses.order_by('id')  # Ordenar si es necesario
            
            num_recipients = recipients.count()
            
            # Determinar el estado global del archivo
            statuses = recipients.values_list('status', flat=True)
            unique_statuses = set(statuses)
            if len(unique_statuses) == 1 and 'Pendiente' not in unique_statuses:
                file_status = unique_statuses.pop()
            else:
                file_status = 'Pendiente'
            
            # Obtener el color de fondo para el archivo
            fill_color_file = status_colors.get(file_status, 'FFFFFF')  # Default White
            
            # Definir el color de fondo para el estado de cada destinatario
            def get_fill_color(status):
                return PatternFill(start_color=status_colors.get(status, 'F2F2F2'),
                                   end_color=status_colors.get(status, 'F2F2F2'),
                                   fill_type="solid")
            
            # Establecer bordes finos
            thin_border = Border(
                left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000')
            )
            
            for idx, cs in enumerate(recipients, start=0):
                row = current_row + idx
                if idx == 0:
                    # Nombre del Archivo solo en la primera fila
                    cell_a = ws.cell(row=row, column=1, value=file.original_name)
                    cell_a.font = Font(name='Calibri', bold=True)
                    cell_a.alignment = Alignment(horizontal='left', vertical='center')
                    cell_a.fill = PatternFill(start_color=fill_color_file, end_color=fill_color_file, fill_type="solid")
                    cell_a.border = thin_border
                # Datos del destinatario
                cell_b = ws.cell(row=row, column=2, value=cs.contact.name)
                cell_c = ws.cell(row=row, column=3, value=cs.contact.email)
                cell_d = ws.cell(row=row, column=4, value=cs.contact.phone)
                cell_e = ws.cell(row=row, column=5, value=cs.status)
                
                # Aplicar fuente Calibri y alineación centrada
                for cell in [cell_b, cell_c, cell_d, cell_e]:
                    cell.font = Font(name='Calibri')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                
                # Aplicar color de fondo basado en el estado
                cell_e.fill = get_fill_color(cs.status)
                
                # Aplicar color de fondo a las columnas de destinatarios
                cell_b.fill = get_fill_color(cs.status)
                cell_c.fill = get_fill_color(cs.status)
                cell_d.fill = get_fill_color(cs.status)
            
            # Si hay múltiples destinatarios, fusionar las celdas de la columna A
            if num_recipients > 1:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + num_recipients -1, end_column=1)
            
            current_row += num_recipients  # Avanzar según el número de destinatarios
        
        # Ajustar el ancho de las columnas
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter  # Obtener la letra de la columna
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Establecer la fuente predeterminada para toda la hoja (Calibri)
        for row in ws.iter_rows():
            for cell in row:
                if not cell.font.name:
                    cell.font = Font(name='Calibri')
                if not cell.alignment.horizontal:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Crear una respuesta HttpResponse con el archivo Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename=estudio_{study.expediente}.xlsx'
        
        wb.save(response)
        return response
    
    except Exception as e:
        return HttpResponse(f"Ocurrió un error al exportar a Excel: {str(e)}", status=500)

@login_required
@user_is_approved
def print_study_pdf(request, study_id):
    # Obtener el estudio asegurando que pertenece al usuario
    study = get_object_or_404(Study, id=study_id, user=request.user)
    
    # Obtener los datos necesarios
    processed_files = ProcessedFile.objects.filter(study=study).order_by('original_name')
    
    correos_enviados = {}
    for file in processed_files:
        contact_statuses = ProcessedFileContactStatus.objects.filter(processed_file=file).select_related('contact').order_by('id')
        if contact_statuses.exists():
            recipients = []
            for cs in contact_statuses:
                recipients.append({
                    'name': cs.contact.name,
                    'email': cs.contact.email,
                    'phone': cs.contact.phone,
                    'status': cs.status,
                })
            recipient_count = len(recipients)
            reclamados = sum(1 for r in recipients if r['status'] == 'Reclamado')
            recibidos = sum(1 for r in recipients if r['status'] == 'Recibido')
            rechazados = sum(1 for r in recipients if r['status'] == 'Rechazado')
            
            if reclamados == recipient_count and recipient_count > 0:
                file_status = 'Reclamado'
            elif recibidos == recipient_count and recipient_count > 0:
                file_status = 'Recibido'
            elif rechazados == recipient_count and recipient_count > 0:
                file_status = 'Rechazado'
            else:
                file_status = 'Pendiente'
            
            correos_enviados[file.id] = {
                'file': file,
                'recipients': recipients,
                'file_status': file_status
            }
    
    context = {
        'study': study,
        'correos_enviados': correos_enviados,
        'estado': study.estado  # Asegúrate de pasar 'estado' si se usa en la plantilla
    }
    
    # Renderizar el HTML
    html_string = render_to_string('study_details_pdf.html', context)
    
    # Crear el PDF
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    pdf = html.write_pdf()
    
    # Crear la respuesta
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=estudio_{study.expediente}.pdf'
    
    return response
    
@login_required
@user_is_approved
def print_selected_studies_pdf(request):
    if request.method == 'POST':
        selected_studies_ids = request.POST.getlist('selected_studies')
        if not selected_studies_ids:
            messages.error(request, "No has seleccionado ningún estudio para imprimir.")
            return redirect('studies')  # Reemplaza 'studies' con el nombre correcto de tu vista de estudios
        
        # Obtener los estudios seleccionados, ordenados alfabéticamente por 'project_name'
        studies = Study.objects.filter(id__in=selected_studies_ids).order_by('project_name').prefetch_related(
            'processedfile_set__processedfilecontactstatus_set__contact'
        )
        
        if not studies.exists():
            messages.error(request, "No se encontraron estudios con los IDs proporcionados.")
            return redirect('studies')
        
        # Preparar los datos estructurados para el PDF
        studies_data = []
        for study in studies:
            # Obtener los archivos ordenados alfabéticamente por 'original_name'
            files = ProcessedFile.objects.filter(study=study).order_by('original_name').prefetch_related(
                'processedfilecontactstatus_set__contact'
            )
            files_data = []
            for file in files:
                # Obtener los destinatarios ordenados alfabéticamente por 'contact__name'
                recipients = file.processedfilecontactstatus_set.select_related('contact').order_by('contact__name')
                recipients_data = []
                for cs in recipients:
                    recipients_data.append({
                        'name': cs.contact.name,
                        'email': cs.contact.email,
                        'phone': cs.contact.phone,
                        'status': cs.status,
                    })
                
                # Determinar el estado global del archivo
                statuses = set(cs.status for cs in recipients)
                if len(statuses) == 1:
                    file_status = statuses.pop()
                else:
                    file_status = 'Pendiente'
                
                files_data.append({
                    'original_name': file.original_name,
                    'recipients': recipients_data,
                    'file_status': file_status,
                    'recipient_count': len(recipients_data),
                })
            
            studies_data.append({
                'expediente': study.expediente,
                'project_name': study.project_name,
                'due_date': study.due_date,
                'estado': study.estado,
                'files': files_data,
            })
        
        context = {
            'studies': studies_data,
        }
        # Renderizar el HTML que incluye todos los estudios con saltos de página
        html_string = render_to_string('multiple_studies_pdf.html', context)
        # Generar el PDF usando WeasyPrint
        html = HTML(string=html_string, base_url=request.build_absolute_uri())
        pdf_bytes = html.write_pdf()
        
        # Crear la respuesta HttpResponse
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=detalles_estudios_seleccionados.pdf'
        
        messages.success(request, "El PDF de los estudios seleccionados ha sido generado exitosamente.")
        return response
    else:
        messages.error(request, "Método no permitido.")
        return redirect('studies')

    
def get_correo_enviados(study):
    """
    Obtiene los correos enviados para un estudio específico y determina el estado global de cada archivo.
    """
    correos_enviados = {}
    # Ordenar los archivos alfabéticamente por 'original_name'
    processed_files = ProcessedFile.objects.filter(study=study).order_by('original_name')
    for file in processed_files:
        # Obtener los estados de contacto ordenados por nombre de contacto
        contact_statuses = ProcessedFileContactStatus.objects.filter(processed_file=file).select_related('contact').order_by('contact__name')
        if contact_statuses.exists():
            recipients = []
            for cs in contact_statuses:
                recipients.append({
                    'name': cs.contact.name,
                    'email': cs.contact.email,
                    'phone': cs.contact.phone,
                    'status': cs.status,
                })
            recipient_count = len(recipients)
            reclamados = sum(1 for r in recipients if r['status'] == 'Reclamado')
            recibidos = sum(1 for r in recipients if r['status'] == 'Recibido')
            rechazados = sum(1 for r in recipients if r['status'] == 'Rechazado')
            
            # Determinar el estado global del archivo
            if reclamados == recipient_count and recipient_count > 0:
                file_status = 'Reclamado'
            elif recibidos == recipient_count and recipient_count > 0:
                file_status = 'Recibido'
            elif rechazados == recipient_count and recipient_count > 0:
                file_status = 'Rechazado'
            else:
                file_status = 'Pendiente'
            
            correos_enviados[file.id] = {
                'file': file,
                'recipients': recipients,
                'file_status': file_status
            }
    return correos_enviados

@login_required
@user_is_approved
def delete_processed_file(request, file_id):
    # Obtener el archivo
    file = get_object_or_404(ProcessedFile, id=file_id, user=request.user)

    if request.method == 'POST':
        # Borrar el archivo del sistema de archivos
        file_path = file.file.path
        if os.path.exists(file_path):
            os.remove(file_path)
        # Borrar el registro de la base de datos
        file.delete()
        messages.success(request, 'Archivo eliminado exitosamente.')
        # Redirigir a la página de upload con el estudio seleccionado
        return redirect(f"{reverse('upload_file')}?study={file.study.id}")
    else:
        return HttpResponseForbidden()

@login_required
@user_is_approved
def delete_processing_session(request, session_id):
    # Obtener la sesión
    session = get_object_or_404(ProcessingSession, id=session_id, user=request.user)

    if request.method == 'POST':
        # Borrar todos los archivos asociados a la sesión del sistema de archivos
        for file in session.processed_files.all():
            file_path = file.file.path
            if os.path.exists(file_path):
                os.remove(file_path)
        # Borrar la carpeta de la sesión
        expediente_folder = os.path.join(os.getcwd(), 'media', 'processed_files', str(request.user.id), str(session.study.id), str(session.id))
        if os.path.exists(expediente_folder):
            try:
                os.rmdir(expediente_folder)
            except OSError:
                # La carpeta no está vacía, eliminar recursivamente
                import shutil
                shutil.rmtree(expediente_folder)
        # Borrar la sesión y los archivos asociados de la base de datos
        session.delete()
        messages.success(request, 'Sesión y todos sus archivos han sido eliminados exitosamente.')
        # Redirigir a la página de upload con el estudio seleccionado
        return redirect(f"{reverse('upload_file')}?study={session.study.id}")
    else:
        return HttpResponseForbidden()
    
@login_required
def approve_users(request):
    user_profile = UserProfile.objects.get(user=request.user)
    if user_profile.role not in ['ADMIN_TOTAL', 'ADMIN_DEPARTAMENTO']:
        messages.error(request, 'No tienes permiso para acceder a esta página.')
        return redirect('home')

    # Filtrar usuarios pendientes según el rol del aprobador
    if user_profile.role == 'ADMIN_TOTAL':
        pending_users = UserProfile.objects.filter(company=user_profile.company, is_approved=False)
    else:
        pending_users = UserProfile.objects.filter(company=user_profile.company, department=user_profile.department, is_approved=False)

    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        action = request.POST.get('action')
        target_profile = UserProfile.objects.get(id=user_id, company=user_profile.company)

        if action == 'approve':
            target_profile.is_approved = True
            target_profile.save()
            messages.success(request, f"Usuario {target_profile.user.username} aprobado.")
        elif action == 'reject':
            target_profile.user.delete()
            messages.success(request, f"Usuario {target_profile.user.username} rechazado y eliminado.")

        return redirect('approve_users')

    return render(request, 'accounts/approve_users.html', {'pending_users': pending_users})

@admin_required
def admin_dashboard(request):
    user_profile = request.user.userprofile
    company = user_profile.company

    # Determinar si el usuario es Administrador Total o Administrador de Departamento
    is_admin_total = user_profile.role == 'ADMIN_TOTAL'
    is_admin_departamento = user_profile.role == 'ADMIN_DEPARTAMENTO'

    if request.method == 'POST':
        # Crear Departamento (solo Administrador Total)
        if 'create_department' in request.POST and is_admin_total:
            dept_form = DepartmentForm(request.POST)
            if dept_form.is_valid():
                department = dept_form.save(commit=False)
                department.company = company
                department.save()
                messages.success(request, f"Departamento '{department.name}' creado exitosamente.")
                return redirect('admin_dashboard')
            else:
                messages.error(request, "Error al crear el departamento. Verifique los datos ingresados.")

        # Editar Departamento (solo Administrador Total)
        elif 'edit_department' in request.POST and is_admin_total:
            department_id = request.POST.get('department_id')
            department = get_object_or_404(Department, id=department_id, company=company)
            dept_form = DepartmentForm(request.POST, instance=department)
            if dept_form.is_valid():
                dept_form.save()
                messages.success(request, f"Departamento '{department.name}' actualizado exitosamente.")
                return redirect('admin_dashboard')
            else:
                messages.error(request, "Error al actualizar el departamento. Verifique los datos ingresados.")

        # Eliminar Departamento (solo Administrador Total)
        elif 'delete_department' in request.POST and is_admin_total:
            department_id = request.POST.get('department_id')
            department = get_object_or_404(Department, id=department_id, company=company)
            if department.name == "Todos":
                messages.error(request, "No puedes eliminar el departamento 'Todos'.")
            else:
                department.delete()
                messages.success(request, f"Departamento '{department.name}' eliminado exitosamente junto con sus usuarios.")
            return redirect('admin_dashboard')

        # Aprobar/Rechazar Usuarios
        elif 'approve_user' in request.POST:
            user_id = request.POST.get('user_id')
            action = request.POST.get('action')
            try:
                user_profile_to_update = UserProfile.objects.get(user__id=user_id, company=company)
                # Verificar permisos
                if is_admin_departamento and user_profile_to_update.department != user_profile.department:
                    messages.error(request, "No tienes permisos para aprobar usuarios de otros departamentos.")
                else:
                    if action == 'approve':
                        user_profile_to_update.is_approved = True
                        user_profile_to_update.save()
                        messages.success(request, f"Usuario '{user_profile_to_update.user.username}' aprobado exitosamente.")
                    elif action == 'decline':
                        user = user_profile_to_update.user
                        user.delete()
                        messages.success(request, f"Usuario '{user.username}' ha sido eliminado.")
            except UserProfile.DoesNotExist:
                messages.error(request, "Usuario no encontrado.")
            return redirect('admin_dashboard')

        # Eliminar Usuario (Administradores Totales y Administradores de Departamento)
        elif 'delete_user' in request.POST:
            user_id = request.POST.get('delete_user_id')
            try:
                user_profile_to_delete = UserProfile.objects.get(user__id=user_id, company=company)
                # Administradores de Departamento no pueden eliminar Administradores Totales
                if is_admin_departamento and user_profile_to_delete.role == 'ADMIN_TOTAL':
                    messages.error(request, "No tienes permisos para eliminar Administradores Totales.")
                elif is_admin_departamento and user_profile_to_delete.department != user_profile.department:
                    messages.error(request, "No tienes permisos para eliminar usuarios de otros departamentos.")
                else:
                    user_to_delete = user_profile_to_delete.user
                    user_to_delete.delete()
                    messages.success(request, f"Usuario '{user_to_delete.username}' eliminado exitosamente.")
            except UserProfile.DoesNotExist:
                messages.error(request, "Usuario no encontrado.")
            return redirect('admin_dashboard')

        # Editar Empresa (solo Administrador Total)
        elif 'edit_company' in request.POST and is_admin_total:
            company_form = CompanyForm(request.POST, instance=company)
            if company_form.is_valid():
                company_form.save()
                messages.success(request, "Datos de la empresa actualizados exitosamente.")
                return redirect('admin_dashboard')
            else:
                messages.error(request, "Error al actualizar los datos de la empresa. Verifique los datos ingresados.")

        # Modificar Rol y Departamento de Usuarios (solo Administrador Total)
        elif 'modify_user' in request.POST and is_admin_total:
            user_id = request.POST.get('modify_user_id')
            new_role = request.POST.get('new_role')
            new_department_id = request.POST.get('new_department')
            try:
                user_profile_to_modify = UserProfile.objects.get(user__id=user_id, company=company)
                # No permitir modificar el rol de Administradores Totales
                if user_profile_to_modify.role == 'ADMIN_TOTAL':
                    messages.error(request, "No puedes modificar el rol de otro Administrador Total.")
                else:
                    if new_role:
                        user_profile_to_modify.role = new_role
                        if new_role in ['ADMIN_DEPARTAMENTO', 'TECNICO_DEPARTAMENTO']:
                            if new_department_id:
                                new_department = get_object_or_404(Department, id=new_department_id, company=company)
                                user_profile_to_modify.department = new_department
                            else:
                                messages.error(request, "Debe seleccionar un departamento para este rol.")
                                return redirect('admin_dashboard')
                        else:
                            user_profile_to_modify.department = None
                        user_profile_to_modify.save()
                        messages.success(request, f"Rol y departamento de '{user_profile_to_modify.user.username}' actualizados exitosamente.")
            except UserProfile.DoesNotExist:
                messages.error(request, "Usuario no encontrado.")
            return redirect('admin_dashboard')

    else:
        dept_form = DepartmentForm() if is_admin_total else None
        company_form = CompanyForm(instance=company) if is_admin_total else None

    # Aprobar Usuarios (solo los de su departamento si es Administrador de Departamento)
    pending_users = UserProfile.objects.filter(is_approved=False, company=company)
    if is_admin_departamento:
        pending_users = pending_users.filter(department=user_profile.department)

    # Usuarios para la sección "Usuarios" (solo aprobados)
    if is_admin_total:
        users = UserProfile.objects.filter(company=company, is_approved=True).exclude(role='ADMIN_TOTAL')
    elif is_admin_departamento:
        users = UserProfile.objects.filter(company=company, department=user_profile.department, is_approved=True).exclude(role='ADMIN_TOTAL')
    else:
        users = UserProfile.objects.none()

    # Obtener todos los departamentos para los formularios de modificación de usuario
    departments = Department.objects.filter(company=company)

    context = {
        'dept_form': dept_form,
        'pending_users': pending_users,
        'company': company,
        'company_form': company_form,
        'users': users,
        'departments': departments,
        'is_admin_total': is_admin_total,
        'is_admin_departamento': is_admin_departamento,
    }

    return render(request, 'accounts/admin_dashboard.html', context)

